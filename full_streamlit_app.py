import pandas as pd
import numpy as np
import re
from datetime import datetime
import os
from typing import Dict, List, Tuple, Optional
import logging
from openpyxl import Workbook
from dateutil import parser
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# Configuration class for customizing conversion behavior
class ConversionConfig:
    """Configuration settings for Form 3 to Circular 29 conversion"""

    def __init__(self):
        # Section keywords (case-insensitive)
        self.section_keywords = {
            'sold': ['sold', 'booked'],
            'unsold': ['unsold', 'available'],
            'tenant': ['tenant', 'rented'],
            'landowner': ['landowner', 'land owner', 'owner'],
            'rehab': ['existing', 'members', 'member'],
            'cidco': ['CIDCO', 'NMMC'],
            'pap': ['PAP']
        }

        # Column mapping keywords
        self.column_keywords = {
            'sr_no': ['sr', 'serial', 'no', 'number'],
            'flat_no': ['flat', 'unit', 'shop', 'apartment'],
            'carpet_area': ['carpet', 'area', 'sq', 'mtrs'],
            'unit_type': ['type', 'category'],
            'building_no': ['building', 'wing', 'tower', 'block']
        }

        # Sheet name patterns
        self.sheet_patterns = {
            'table_a': ['table a'],
            'table_b': ['table b'],
            'table_c': ['table c']
        }

        # Output formatting
        self.output_config = {
            'start_row': 11,
            'column_widths': [8, 15, 20, 18, 35, 25],
            'font_size': 11,
            'header_font_size': 12
        }

        # Regex patterns
        self.regex_patterns = {
            'project_info': r'for the project\s+(.+?)\s+having',
            'rera_number': r'having maharera registration number\s+([A-Z0-9]+)\s+being developed',
            'date_pattern': r'\(as on\s+(.+?)\)',
            'filename_rera': r'([A-Z]\d+)',
            'filename_date': r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})'
        }


class Form3ToCircular29Converter:
    """
    Converts Form 3 Excel files to Circular 29 format
    """

    def __init__(self, config: ConversionConfig = None):
        self.include_building_column = False  # Will be set during unit extractio
        self.config = config or ConversionConfig()
        self.project_name = ""
        self.rera_number = ""
        self.as_on_date = ""
        self.unit_sections = {
            'sold': [],
            'unsold': [],
            'tenant': [],
            'landowner': [],
            'rehab': [],
            'cidco': [],
            'pap': []
        }

    def normalize_text(self, text: str) -> str:
        """Normalize text by removing extra whitespace and converting to lowercase"""
        if not isinstance(text, str):
            return str(text).strip().lower()
        return text.strip().lower()

    def format_date(self, date_str: str) -> str:
        """
        Convert various date formats (like '30th June 2025') into 'DD/MM/YYYY'
        """
        try:
            parsed_date = parser.parse(date_str, fuzzy=True)
            return parsed_date.strftime("%d/%m/%Y")
        except Exception as e:
            logger.warning(f"Unable to parse as-on date: {date_str} — {e}")
            return date_str  # fallback to original

    def extract_project_info(self, sheet_data: pd.DataFrame) -> Tuple[str, str]:
        """
        Extract project name and RERA number from Table A sheet
        """
        logger.info("Extracting project information from Table A")

        # Convert all cells to string and search for the pattern
        for index, row in sheet_data.iterrows():
            for col in sheet_data.columns:
                cell_value = str(row[col]) if pd.notna(row[col]) else ""

                # Look for the certificate sentence pattern
                if "certificate is being issued for the project" in cell_value.lower():
                    logger.info(f"Found certificate sentence: {cell_value}")

                    # Extract project name
                    project_match = re.search(r'for the project\s+(.+?)\s+having', cell_value, re.IGNORECASE)
                    if project_match:
                        project_name = project_match.group(1).strip()
                        logger.info(f"Extracted project name: {project_name}")

                    # Extract RERA number
                    rera_match = re.search(r'having maharera registration number\s+([A-Z0-9]+)\s+being developed',
                                           cell_value, re.IGNORECASE)
                    if rera_match:
                        rera_number = rera_match.group(1).strip()
                        logger.info(f"Extracted RERA number: {rera_number}")

                    return project_name, rera_number

        logger.warning("Project information not found in Table A")
        return "", ""

    def extract_as_on_date(self, sheet_data: pd.DataFrame) -> str:
        """
        Extract as-on date from Table B sheet
        """
        logger.info("Extracting as-on date from Table B")

        for index, row in sheet_data.iterrows():
            for col in sheet_data.columns:
                cell_value = str(row[col]) if pd.notna(row[col]) else ""

                # Look for the date pattern in Table B
                if "table b" in cell_value.lower() and "as on" in cell_value.lower():
                    logger.info(f"Found date sentence: {cell_value}")

                    # Extract date from parentheses
                    date_match = re.search(r'\(as on\s+(.+?)\)', cell_value, re.IGNORECASE)
                    if date_match:
                        as_on_date = date_match.group(1).strip()
                        logger.info(f"Extracted as-on date: {as_on_date}")
                        return as_on_date

        logger.warning("As-on date not found in Table B")
        return ""

    def find_section_start(self, sheet_data: pd.DataFrame, section_keyword: str) -> int:
        logger.info(f"Looking for section: {section_keyword}")
        for index, row in sheet_data.iterrows():
            for col in sheet_data.columns:
                cell_value = str(row[col]) if pd.notna(row[col]) else ""
                if section_keyword.lower() in cell_value.lower():
                    logger.info(f"Found section '{section_keyword}' at row {index}")
                    return index

        logger.warning(f"Section '{section_keyword}' not found")
        return -1


    def find_data_start_row(self, sheet_data: pd.DataFrame, start_row: int) -> int:
        """
        Find the row where Sr. No equals exactly 1
        """
        for index in range(start_row, len(sheet_data)):
            row = sheet_data.iloc[index]
            for col in sheet_data.columns:
                cell_value = row[col]
                # Check if Sr. No is exactly 1 (not "1." or "(1)")
                if cell_value == 1 or cell_value == "1":
                    logger.info(f"Found Sr. No = 1 at row {index}")
                    return index

        logger.warning("Sr. No = 1 not found")
        return -1

    def extract_from_filename(self, filename: str) -> Tuple[str, str, str]:
        """
        Extract project name, RERA number, and date from filename as fallback
        """
        logger.info(f"Extracting information from filename: {filename}")

        # Remove file extension
        base_name = os.path.splitext(filename)[0]

        # Try to extract RERA number pattern
        rera_match = re.search(r'([A-Z]\d+)', base_name)
        rera_number = rera_match.group(1) if rera_match else ""

        # Try to extract date pattern
        date_match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', base_name)
        date_str = date_match.group(1) if date_match else ""

        # Extract project name (everything before RERA number or date)
        project_name = base_name
        if rera_match:
            project_name = base_name[:rera_match.start()].strip()
        elif date_match:
            project_name = base_name[:date_match.start()].strip()

        # Clean up project name
        project_name = re.sub(r'[_-]+', ' ', project_name).strip()

        logger.info(f"Extracted from filename - Project: {project_name}, RERA: {rera_number}, Date: {date_str}")
        return project_name, rera_number, date_str

    def extract_unit_data(self, sheet_data: pd.DataFrame, section_keyword: str) -> List[Dict]:
        """
        Extract unit data for a specific section
        """
        logger.info(f"Extracting unit data for section: {section_keyword}")

        # Find section start
        section_start = self.find_section_start(sheet_data, section_keyword)
        if section_start == -1:
            return []

        # Find data start (where Sr. No = 1)
        data_start = self.find_data_start_row(sheet_data, section_start)
        if data_start == -1:
            return []

        # Get header row (usually the row before data starts)
        # Try to locate the actual header row by looking for known keywords
        header_keywords = ["sr", "flat", "carpet", "unit", "building", "wing"]
        header_row_index = -1

        for i in range(max(0, data_start - 3), data_start + 1):
            row = sheet_data.iloc[i]
            match_count = 0
            for cell in row:
                text = str(cell).strip().lower()
                if any(keyword in text for keyword in header_keywords):
                    match_count += 1
            if match_count >= 2:
                header_row_index = i
                break

        if header_row_index == -1:
            logger.warning(f"Could not reliably locate header row for section: {section_keyword}")
            return []

        # Extract headers
        headers = []
        header_row = sheet_data.iloc[header_row_index]
        for col in sheet_data.columns:
            header_value = str(header_row[col]) if pd.notna(header_row[col]) else ""
            headers.append(self.normalize_text(header_value))

        logger.info(f"Headers found: {headers}")

        # Map column indices
        col_mapping = {}
        for i, header in enumerate(headers):
            if "sr" in header and "no" in header and 'sr_no' not in col_mapping:
                col_mapping['sr_no'] = i
            elif ("flat" in header and "no" in header or "shop" in header and "no" in header) and 'flat_no' not in col_mapping:
                col_mapping['flat_no'] = i
            elif "carpet" in header and "area" in header and 'carpet_area' not in col_mapping:
                col_mapping['carpet_area'] = i
            elif "unit" in header and "type" in header and "apartment" not in header and 'unit_type' not in col_mapping:
                col_mapping['unit_type'] = i
            elif "building" in header and "no" in header and 'building_no' not in col_mapping:
                col_mapping['building_no'] = i
            elif "wing" in header and 'building_no' not in col_mapping:
                col_mapping['building_no'] = i
            if 'building_no' in col_mapping:
                self.include_building_column = True

        logger.info(f"Column mapping: {col_mapping}")

        # Extract data rows
        units = []
        section_keywords = list(self.config.section_keywords.keys())
        other_section_keywords = [k.lower() for k in section_keywords if k.lower() != section_keyword.lower()]
        parsing_started = False

        for row_index in range(data_start, len(sheet_data)):
            row = sheet_data.iloc[row_index]

            # # 🔴 STOP if any new section header is detected (always, even if parsing hasn't started)
            # row_text = ' '.join(str(cell).lower() for cell in row if pd.notna(cell)).strip()
            # if any(keyword in row_text for keyword in other_section_keywords):
            #     logger.info(
            #         f"Detected new section while scanning '{section_keyword}' → '{row_text}' at row {row_index}. Stopping.")
            #     break

            # ✅ Look for Sr. No = 1 to start parsing
            sr_no_col = col_mapping.get('sr_no', 0)
            sr_no_cell = row.iloc[sr_no_col] if sr_no_col < len(row) else None

            if not parsing_started:
                if str(sr_no_cell).strip() in ["1", 1]:
                    parsing_started = True
                else:
                    continue  # keep looking for Sr. No = 1

            # Once started, stop if Sr. No becomes invalid
            if pd.isna(sr_no_cell) or not str(sr_no_cell).strip():
                break
            try:
                sr_no = int(float(sr_no_cell))
            except (ValueError, TypeError):
                break

            # ✅ Extract unit data
            bldg_index = col_mapping.get('building_no', 1)
            flat_index = col_mapping.get('flat_no', 2)
            carpet_index = col_mapping.get('carpet_area', 3)

            unit_data = {
                'sr_no': sr_no,
                'building_no': str(row.iloc[bldg_index]).strip() if (
                        bldg_index < len(row) and pd.notna(row.iloc[bldg_index])) else "",
                'flat_no': str(row.iloc[flat_index]).strip() if (
                        flat_index < len(row) and pd.notna(row.iloc[flat_index])) else "",
                'carpet_area': str(row.iloc[carpet_index]).strip() if (
                        carpet_index < len(row) and pd.notna(row.iloc[carpet_index])) else "",
                'status': section_keyword.lower(),
                'registration_date': ""
            }

            units.append(unit_data)

        logger.info(f"Extracted {len(units)} units for section '{section_keyword}'")
        return units

    def process_form3_file(self, file_path: str) -> bool:
        """
        Process the Form 3 Excel file and extract all required information
        """
        logger.info(f"Processing Form 3 file: {file_path}")

        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            logger.info(f"Found sheets: {sheet_names}")

            # Find Table A (Project Information)
            table_a_sheet = None
            for sheet_name in sheet_names:
                if "table a" in sheet_name.lower():
                    table_a_sheet = sheet_name
                    break

            if table_a_sheet:
                table_a_data = pd.read_excel(file_path, sheet_name=table_a_sheet, header=None)
                self.project_name, self.rera_number = self.extract_project_info(table_a_data)

            # Find Table B (As-on Date)
            table_b_sheet = None
            for sheet_name in sheet_names:
                if "table b" in sheet_name.lower():
                    table_b_sheet = sheet_name
                    break

            if table_b_sheet:
                table_b_data = pd.read_excel(file_path, sheet_name=table_b_sheet, header=None)
                self.as_on_date = self.extract_as_on_date(table_b_data)

            # Find Table C (Unit Details)
            table_c_sheet = None
            for sheet_name in sheet_names:
                if "table c" in sheet_name.lower():
                    table_c_sheet = sheet_name
                    break

            if table_c_sheet:
                table_c_data = pd.read_excel(file_path, sheet_name=table_c_sheet, header=None, dtype=str)

                # Extract unit data for each section
                sections = ['sold', 'unsold', 'tenant', 'landowner', 'rehab', 'cidco', 'pap']
                for section in sections:
                    units = self.extract_unit_data(table_c_data, section)
                    self.unit_sections[section] = units

            # Fallback to filename if information not found
            if not self.project_name or not self.rera_number or not self.as_on_date:
                filename = os.path.basename(file_path)
                fallback_project, fallback_rera, fallback_date = self.extract_from_filename(filename)

                if not self.project_name:
                    self.project_name = fallback_project
                if not self.rera_number:
                    self.rera_number = fallback_rera
                if not self.as_on_date:
                    self.as_on_date = fallback_date

            logger.info(f"Final extracted data:")
            logger.info(f"Project Name: {self.project_name}")
            logger.info(f"RERA Number: {self.rera_number}")
            logger.info(f"As-on Date: {self.as_on_date}")

            for section, units in self.unit_sections.items():
                logger.info(f"{section.upper()}: {len(units)} units")

            return True

        except Exception as e:
            logger.error(f"Error processing Form 3 file: {str(e)}")
            return False

    def create_circular29_excel(self, output_path: str) -> bool:
        """
        Create the Circular 29 Excel file with proper formatting
        """
        logger.info(f"Creating Circular 29 file: {output_path}")

        try:
            from openpyxl.styles import Border, Side, Alignment, Font

            # Define fonts
            aptos_font_start = Font(name='Aptos', size=14, bold=True)
            aptos_font = Font(name='Aptos', size=11)
            aptos_bold = Font(name='Aptos', bold=True, size=14)
            times_new_roman_font = Font(name='Times New Roman', size=11)
            times_new_roman_bold = Font(name='Times New Roman', bold=True, size=12)

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Set default row height
            ws.sheet_format.defaultRowHeight = 15.6

            # Set column widths
            column_widths = {
                'A': 9,
                'B': 15,
                'C': 12,
                'D': 16,
                'E': 18
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Define styles
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Current row tracker
            current_row = 2

            # Initiate File
            ws[f'A{current_row}'] = f"To whomsoever it may concern"
            ws[f'A{current_row}'].font = aptos_font_start
            current_row += 2

            # Add project information
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = f"Name of Project: {self.project_name}"
            ws[f'A{current_row}'].font = aptos_font_start
            current_row += 1

            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = f"MahaRERA Project Registration Number: {self.rera_number}"
            ws[f'A{current_row}'].font = aptos_font_start
            current_row += 2

            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = "Information of Sold/ Booked inventory (Building Wise)"
            ws[f'A{current_row}'].font = aptos_bold
            current_row += 2

            ws.merge_cells(f'A{current_row}:E{current_row}')
            formatted_date = self.format_date(self.as_on_date)
            ws[f'A{current_row}'] = f"Information as on dated: {formatted_date}"
            ws[f'A{current_row}'].font = aptos_font_start
            current_row += 2

            # Check if any unit has non-empty building_no
            include_building_no = self.include_building_column

            # Add table headers conditionally
            headers = ["Sr.No"]
            if include_building_no:
                headers.append("Building No")
            headers += [
                "Flat No./ Shop No",
                "Carpet Area In Sq.Mtrs ",
                "Sold/ Booked /Unsold Reserved/ Rehab/ Mortgaged/ Not for Sale",
                "Registration Date of Sub Registrar"
            ]

            for i, header in enumerate(headers):
                col_letter = chr(65 + i)
                cell = ws[f'{col_letter}{current_row}']
                cell.value = header
                cell.font = times_new_roman_bold
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[current_row].height = 98  # Adjusted for multiline header

            current_row += 1
            unit_counter = 1
            section_order = ['sold', 'unsold', 'landowner', 'tenant', 'rehab', 'cidco', 'pap']

            for section in section_order:
                units = self.unit_sections.get(section, [])
                if not units:
                    continue

                for unit in units:
                    col = 1
                    # Sr.No column
                    ws.cell(row=current_row, column=col, value=unit_counter).font = times_new_roman_font
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=current_row, column=col).border = thin_border
                    col += 1

                    # Building No column (if included)
                    if include_building_no:
                        bldg_val = str(unit.get('building_no', '')).strip()
                        ws.cell(row=current_row, column=col, value=bldg_val)
                        ws.cell(row=current_row, column=col).font = times_new_roman_font
                        ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center",
                                                                                   vertical="center")
                        ws.cell(row=current_row, column=col).border = thin_border
                        col += 1

                    # Flat No./Shop No column - THIS IS THE FIX
                    flat_no_val = str(unit.get('flat_no', '')).strip()
                    ws.cell(row=current_row, column=col, value=flat_no_val)
                    ws.cell(row=current_row, column=col).font = times_new_roman_font
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=current_row, column=col).border = thin_border
                    col += 1

                    # Carpet Area column
                    carpet_val = unit.get('carpet_area', '')
                    try:
                        carpet_val = f"{float(carpet_val):.2f}"
                    except:
                        pass
                    ws.cell(row=current_row, column=col, value=carpet_val).font = times_new_roman_font
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=current_row, column=col).border = thin_border
                    col += 1

                    # Status column
                    if section in ['cidco', 'pap']:
                        status_display = section.upper()
                    else:
                        status_display = section.capitalize()
                    ws.cell(row=current_row, column=col, value=status_display).font = times_new_roman_font
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=current_row, column=col).border = thin_border
                    col += 1

                    # Registration Date column
                    if section == 'sold':
                        reg_date_val = ""  # ✅ MODIFIED: Show 'NA' for unsold units
                    else:
                        reg_date_val = "NA"  # Or you can replace with actual date if available

                    ws.cell(row=current_row, column=col, value=reg_date_val).font = times_new_roman_font
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=current_row, column=col).border = thin_border

                    unit_counter += 1
                    current_row += 1

            # Add only one blank row before Note
            current_row += 1

            note_text = "Note: This information has been tallied and confirmed from details submitted in Annexure 'A' of Form 3 issued by Chartered Accountant."

            # Merge A to E for the note
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = note_text
            ws[f'A{current_row}'].font = aptos_font
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # Set height of note row
            ws.row_dimensions[current_row].height = 38

            current_row += 2

            # Merge E:F for "Sign"
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=4, value="Sign").font = aptos_font
            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="center")
            current_row += 1

            ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=5, value="________________").font = aptos_font
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal="center")

            wb.save(output_path)
            logger.info(f"Circular 29 file created successfully: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error creating Circular 29 file: {str(e)}")
            return False


def main():
    """
    Main function to run the converter
    """
    # Example usage
    input_file = r"D:\Projects\Circular29-automate\resources\Form 3 - Swarg - June 2025.xlsx"
    output_dir = r"D:\Projects\Circular29-automate\saved data"

    # Create converter instance
    converter = Form3ToCircular29Converter()

    # Process the Form 3 file
    if converter.process_form3_file(input_file):
        # Format project name
        project_clean = converter.project_name.strip().replace(" ", "_")

        # Format as-on date for filename (e.g., "June 2025")
        try:
            parsed_date = datetime.strptime(converter.format_date(converter.as_on_date), "%d/%m/%Y")
            as_on_string = parsed_date.strftime("%B %Y")
        except:
            as_on_string = converter.as_on_date.replace("/", " ").replace("-", " ")

        # Final output path
        output_filename = f"Circular 29 - {project_clean.replace('_', ' ')} as on {as_on_string}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        # Create Circular 29
        if converter.create_circular29_excel(output_path):
            print("✅ Conversion completed successfully!")
            print(f"📁 Output file: {output_path}")
        else:
            print("❌ Error creating Circular 29 file")
    else:
        print("❌ Error processing Form 3 file")


if __name__ == "__main__":
    main()


# Example usage function for different scenarios
def convert_form3_to_circular29(input_file_path: str, output_file_path: str = None,
                                config: ConversionConfig = None) -> bool:
    """
    Convenience function to convert Form 3 to Circular 29

    Args:
        input_file_path: Path to the Form 3 Excel file
        output_file_path: Path for the output Circular 29 file (optional)
        config: Custom configuration object (optional)

    Returns:
        bool: True if conversion successful, False otherwise
    """
    if not output_file_path:
        base_name = os.path.splitext(input_file_path)[0]
        output_file_path = f"{base_name}_Circular29.xlsx"

    converter = Form3ToCircular29Converter(config)

    if converter.process_form3_file(input_file_path):
        return converter.create_circular29_excel(output_file_path)

    return False


# Validation and Quality Control Functions
def validate_form3_file(file_path: str) -> Dict[str, bool]:
    """
    Validate if a file appears to be a valid Form 3 Excel file

    Args:
        file_path: Path to the Excel file

    Returns:
        dict: Validation results with details
    """
    validation_results = {
        'is_excel': False,
        'has_table_a': False,
        'has_table_b': False,
        'has_table_c': False,
        'has_project_info': False,
        'has_date_info': False,
        'has_unit_data': False,
        'overall_valid': False
    }

    try:
        # Check if file is Excel
        if not file_path.endswith(('.xlsx', '.xls')):
            return validation_results

        validation_results['is_excel'] = True

        # Read Excel file
        excel_file = pd.ExcelFile(file_path)
        sheet_names = [name.lower() for name in excel_file.sheet_names]

        # Check for required sheets
        validation_results['has_table_a'] = any('table a' in name for name in sheet_names)
        validation_results['has_table_b'] = any('table b' in name for name in sheet_names)
        validation_results['has_table_c'] = any('table c' in name for name in sheet_names)

        # Quick content validation
        if validation_results['has_table_a']:
            table_a = pd.read_excel(file_path, sheet_name=excel_file.sheet_names[0], header=None)
            validation_results['has_project_info'] = any(
                'certificate is being issued for the project' in str(cell).lower()
                for row in table_a.values for cell in row if pd.notna(cell)
            )

        if validation_results['has_table_b']:
            table_b_name = next(name for name in excel_file.sheet_names if 'table b' in name.lower())
            table_b = pd.read_excel(file_path, sheet_name=table_b_name, header=None)
            validation_results['has_date_info'] = any(
                'as on' in str(cell).lower()
                for row in table_b.values for cell in row if pd.notna(cell)
            )

        if validation_results['has_table_c']:
            table_c_name = next(name for name in excel_file.sheet_names if 'table c' in name.lower())
            table_c = pd.read_excel(file_path, sheet_name=table_c_name, header=None)
            validation_results['has_unit_data'] = any(
                any(keyword in str(cell).lower() for keyword in ['sold', 'unsold', 'tenant', 'landowner', 'rehab', 'cidco', 'pap'])
                for row in table_c.values for cell in row if pd.notna(cell)
            )

        # Overall validation
        validation_results['overall_valid'] = all([
            validation_results['is_excel'],
            validation_results['has_table_a'] or validation_results['has_table_b'] or validation_results['has_table_c'],
            validation_results['has_project_info'] or validation_results['has_date_info'] or validation_results[
                'has_unit_data']
        ])

    except Exception as e:
        logger.error(f"Validation error: {str(e)}")

    return validation_results


def generate_conversion_report(converter: Form3ToCircular29Converter, input_file: str, output_file: str) -> str:
    """
    Generate a detailed conversion report

    Args:
        converter: Form3ToCircular29Converter instance
        input_file: Input file path
        output_file: Output file path

    Returns:
        str: Formatted conversion report
    """
    report = f"""
📊 FORM 3 TO CIRCULAR 29 CONVERSION REPORT
{'=' * 50}

📁 INPUT FILE: {input_file}
📁 OUTPUT FILE: {output_file}
📅 CONVERSION DATE: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

🏢 PROJECT INFORMATION:
   • Project Name: {converter.project_name or 'Not found'}
   • RERA Number: {converter.rera_number or 'Not found'}
   • As-on Date: {converter.as_on_date or 'Not found'}

📋 UNIT SUMMARY:
"""

    total_units = 0
    for section, units in converter.unit_sections.items():
        count = len(units)
        total_units += count
        report += f"   • {section.upper()}: {count} units\n"

    report += f"   • TOTAL UNITS: {total_units}\n"

    if total_units > 0:
        report += f"\n✅ CONVERSION STATUS: SUCCESS"
    else:
        report += f"\n❌ CONVERSION STATUS: NO UNITS FOUND"

    return report


# Error Recovery Functions
def attempt_data_recovery(file_path: str, converter: Form3ToCircular29Converter) -> bool:
    """
    Attempt to recover data using alternative methods

    Args:
        file_path: Path to the Form 3 file
        converter: Converter instance

    Returns:
        bool: True if recovery successful
    """
    logger.info("Attempting data recovery...")

    try:
        # Try reading all sheets without specific sheet names
        excel_file = pd.ExcelFile(file_path)

        for sheet_name in excel_file.sheet_names:
            logger.info(f"Attempting recovery from sheet: {sheet_name}")

            try:
                sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

                # Try to find project info in any sheet
                if not converter.project_name or not converter.rera_number:
                    project_name, rera_number = converter.extract_project_info(sheet_data)
                    if project_name:
                        converter.project_name = project_name
                    if rera_number:
                        converter.rera_number = rera_number

                # Try to find date info in any sheet
                if not converter.as_on_date:
                    date_info = converter.extract_as_on_date(sheet_data)
                    if date_info:
                        converter.as_on_date = date_info

                # Try to find unit data in any sheet
                sections = ['sold', 'unsold', 'tenant', 'landowner', 'rehab', 'cidco', 'pap']
                for section in sections:
                    if not converter.unit_sections[section]:
                        units = converter.extract_unit_data(sheet_data, section)
                        if units:
                            converter.unit_sections[section] = units

            except Exception as e:
                logger.warning(f"Could not process sheet {sheet_name}: {str(e)}")
                continue

        # Check if we recovered any meaningful data
        has_basic_info = bool(converter.project_name or converter.rera_number)
        has_units = any(converter.unit_sections.values())

        if has_basic_info or has_units:
            logger.info("Data recovery partially successful")
            return True
        else:
            logger.warning("Data recovery failed")
            return False

    except Exception as e:
        logger.error(f"Data recovery error: {str(e)}")
        return False


# Batch processing function
def batch_convert_form3_files(input_directory: str, output_directory: str = None,
                              generate_reports: bool = True) -> Dict[str, bool]:
    """
    Convert multiple Form 3 files to Circular 29 format

    Args:
        input_directory: Directory containing Form 3 Excel files
        output_directory: Directory to save Circular 29 files (optional)
        generate_reports: Whether to generate conversion reports (optional)

    Returns:
        dict: Dictionary with filename as key and success status as value
    """
    if not output_directory:
        output_directory = input_directory

    results = {}
    reports_dir = os.path.join(output_directory, 'conversion_reports')

    if generate_reports:
        os.makedirs(reports_dir, exist_ok=True)

    # Find all Excel files in input directory
    excel_files = [f for f in os.listdir(input_directory) if f.endswith(('.xlsx', '.xls'))]

    print(f"Found {len(excel_files)} Excel files to process")

    for i, file in enumerate(excel_files, 1):
        input_path = os.path.join(input_directory, file)
        output_filename = f"{os.path.splitext(file)[0]}_Circular29.xlsx"
        output_path = os.path.join(output_directory, output_filename)

        print(f"\n[{i}/{len(excel_files)}] Processing: {file}")

        # Validate file first
        validation = validate_form3_file(input_path)
        if not validation['overall_valid']:
            print(f"⚠️  File validation failed: {file}")
            results[file] = False
            continue

        # Convert file
        converter = Form3ToCircular29Converter()
        success = False

        if converter.process_form3_file(input_path):
            success = converter.create_circular29_excel(output_path)

        # Try recovery if initial conversion failed
        if not success:
            print(f"🔄 Attempting data recovery for: {file}")
            if attempt_data_recovery(input_path, converter):
                success = converter.create_circular29_excel(output_path)

        results[file] = success

        if success:
            print(f"✅ Successfully converted: {file}")

            # Generate report if requested
            if generate_reports:
                report = generate_conversion_report(converter, input_path, output_path)
                report_filename = f"{os.path.splitext(file)[0]}_conversion_report.txt"
                report_path = os.path.join(reports_dir, report_filename)

                with open(report_path, 'w', encoding='utf-8') as f:
                    f.write(report)

                print(f"📄 Report generated: {report_filename}")
        else:
            print(f"❌ Failed to convert: {file}")

    # Generate summary report
    if generate_reports:
        summary = generate_batch_summary(results, input_directory, output_directory)
        summary_path = os.path.join(reports_dir, 'batch_conversion_summary.txt')
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(summary)
        print(f"\n📊 Batch summary saved: {summary_path}")

    return results


def generate_batch_summary(results: Dict[str, bool], input_dir: str, output_dir: str) -> str:
    """Generate a summary report for batch conversion"""
    successful = sum(1 for success in results.values() if success)
    failed = len(results) - successful

    summary = f"""
🔄 BATCH CONVERSION SUMMARY
{'=' * 50}

📁 INPUT DIRECTORY: {input_dir}
📁 OUTPUT DIRECTORY: {output_dir}
📅 CONVERSION DATE: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

📊 RESULTS:
   • Total files processed: {len(results)}
   • Successful conversions: {successful}
   • Failed conversions: {failed}
   • Success rate: {(successful / len(results) * 100):.1f}%

📋 DETAILED RESULTS:
"""

    for filename, success in results.items():
        status = "✅ SUCCESS" if success else "❌ FAILED"
        summary += f"   • {filename}: {status}\n"

    if failed > 0:
        summary += f"\n⚠️  FAILED FILES:\n"
        for filename, success in results.items():
            if not success:
                summary += f"   • {filename}\n"

    return summary


# === Streamlit Interface ===

import streamlit as st
import tempfile
import os
from datetime import datetime

# --- Page Config ---
st.set_page_config(
    page_title="Form 3 ➝ Circular 29 Converter",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# --- Styles ---
st.markdown("""
    <style>
    #MainMenu, header, footer {visibility: hidden;}
    .css-164nlkn {display: none;}  /* GitHub icon */

    html, body, .block-container {
        background-color: #0f1117;
        color: white;
    }

    /* Optimize main container spacing */
    .main .block-container {
        padding-top: 0.5rem;
        padding-bottom: 5rem; /* Space for footer */
        max-width: 100% !important;
        min-height: calc(100vh - 5rem);
    }

    /* Ensure the app content area takes full height */
    .stApp {
        min-height: 100vh;
        position: relative;
    }

    .title-text {
        font-size: 2.2rem;
        font-weight: 700;
        text-align: center;
        margin-bottom: 1rem;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }

    .info-card {
        background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
        padding: 1.5rem;
        border-radius: 0.75rem;
        border: 1px solid #475569;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        margin-bottom: 1rem;
    }

    .upload-area {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        padding: 1.5rem;
        border-radius: 0.75rem;
        border: 1px solid #475569;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
    }

    /* Fixed footer styling */
    .footer {
        position: fixed;
        bottom: 0;
        left: 0;
        right: 0;
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        text-align: center;
        padding: 0.8rem 0;
        font-size: 0.9rem;
        font-weight: 500;
        color: #94a3b8;
        border-top: 1px solid #475569;
        z-index: 1000;
        box-shadow: 0 -4px 6px -1px rgba(0, 0, 0, 0.1);
    }

    /* Prevent footer from covering content */
    .main-content {
        padding-bottom: 4rem;
    }

    /* Custom styling for Streamlit components */
    .stFileUploader > div > div > div > div {
        background-color: #1e293b !important;
        border: 2px dashed #475569 !important;
        border-radius: 0.5rem !important;
    }

    .stSuccess {
        background-color: #065f46 !important;
        border: 1px solid #10b981 !important;
    }

    .stError {
        background-color: #7f1d1d !important;
        border: 1px solid #ef4444 !important;
    }
    </style>
""", unsafe_allow_html=True)

# --- Main Content Container ---
st.markdown('<div class="main-content">', unsafe_allow_html=True)

# --- Title ---
st.markdown('<div class="title-text">📄 Form 3 ➝ Circular 29 Converter</div>', unsafe_allow_html=True)

# --- Layout: Info on Left, Upload on Right ---
col_info, col_upload = st.columns([1.5, 2])

# --- LEFT COLUMN: Tool Info ---
with col_info:
    st.markdown("### 🧾 What this tool does")

    st.markdown("✔️ **Extracts Project Name, RERA No., and As-on Date**")
    st.markdown("📊 **Reads Sold / Unsold / Landowner / Tenant unit data**")
    st.markdown("✅ **Generates Circular 29 Excel file as per MahaRERA format**")

    st.markdown("---")
    st.markdown("**Expected Sheets:**")
    st.markdown("• Table A — Project Info")
    st.markdown("• Table B — As-on Date")
    st.markdown("• Table C — Inventory Details")

# --- RIGHT COLUMN: Upload + Status + Download ---
with col_upload:
    st.markdown("### 📂 Upload Form 3 (.xlsx)")

    uploaded_file = st.file_uploader("Choose Form 3 Excel", type=["xlsx", "xls"], label_visibility="collapsed")

    if uploaded_file:
        with st.spinner("⏳ Processing Form 3 file..."):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(uploaded_file.read())
                tmp_path = tmp.name

            converter = Form3ToCircular29Converter()
            if converter.process_form3_file(tmp_path):
                project_clean = converter.project_name.strip().replace(" ", "_") or "Project"
                try:
                    parsed_date = datetime.strptime(converter.format_date(converter.as_on_date), "%d/%m/%Y")
                    as_on_string = parsed_date.strftime("%B %Y")
                except:
                    as_on_string = converter.as_on_date.replace("/", " ").replace("-", " ")

                filename = f"Circular 29 - {project_clean.replace('_', ' ')} as on {as_on_string}.xlsx"
                output_dir = os.path.join("saved data")
                os.makedirs(output_dir, exist_ok=True)
                output_path = os.path.join(output_dir, filename)

                if converter.create_circular29_excel(output_path):
                    st.success("✅ Conversion completed successfully!")
                    with open(output_path, "rb") as f:
                        st.download_button("📥 Download Circular 29 Excel", f.read(), file_name=filename)
                else:
                    st.error("❌ Failed to generate Circular 29 Excel.")
            else:
                st.error("❌ Failed to process Form 3 file.")

# --- Close Main Content Container ---
st.markdown('</div>', unsafe_allow_html=True)

# --- Fixed Footer ---
st.markdown('<div class="footer">© 2025 Aryan Parte. All rights reserved.</div>', unsafe_allow_html=True)
